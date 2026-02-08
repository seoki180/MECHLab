"""
SAE J2951 Report Generator (Template-based)
기존 Excel 템플릿 형식에 맞춰 리포트 생성
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from datetime import datetime
from typing import Dict, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os


def create_radar_chart(er, dr, eer, ascr, iwr, rmsse,
                       title='SAE J2951 Report',
                       save_path=None):
    """
    육각형 레이더 차트 생성

    Parameters:
    - er, dr, eer, ascr, iwr: 백분율 값 (%)
    - rmsse: mph 단위 값
    - save_path: 이미지 저장 경로

    Returns:
    - fig: matplotlib figure 객체
    """

    # 데이터 준비
    categories = ['ER', 'DR', 'EER', 'ASCR', 'IWR', 'RMSSE']
    values = [er, dr, eer, ascr, iwr, rmsse]

    # 육각형 각도 계산
    angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
    values_plot = values + values[:1]
    angles_plot = angles + angles[:1]

    # Figure 생성
    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(projection='polar'))

    # 실제 데이터 플롯 (빨간색)
    ax.plot(angles_plot, values_plot, 'o-', linewidth=2.5,
            color='#DC143C', markersize=8, label='Actual', zorder=3)
    ax.fill(angles_plot, values_plot, alpha=0.3, color='#DC143C', zorder=2)

    # 기준선 (0) - 파란 점선
    zeros = [0] * len(angles_plot)
    ax.plot(angles_plot, zeros, '--', linewidth=2,
            color='#4169E1', alpha=0.7, label='Target', zorder=2)

    # 축 레이블
    ax.set_xticks(angles)
    ax.set_xticklabels(categories, size=13, weight='bold')

    # Y축 범위 및 눈금
    max_val = max(abs(min(values)), abs(max(values)))
    y_limit = max(3, np.ceil(max_val))
    ax.set_ylim(-y_limit, y_limit)

    yticks = list(range(-int(y_limit), int(y_limit) + 1))
    ax.set_yticks(yticks)
    ax.set_yticklabels([str(y) for y in yticks], size=10)

    # 격자선
    ax.grid(True, linestyle='--', alpha=0.4, linewidth=1)

    # 범례
    ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=10)

    # 제목
    plt.title(title, size=15, weight='bold', pad=25)

    plt.tight_layout()

    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight',
                   facecolor='white', edgecolor='none')

    return fig


def create_excel_report_from_template(
    results: Dict[str, float],
    test_info: Dict[str, str],
    excel_path: str,
    chart_path: str = None,
    template_path: str = None
):
    """
    템플릿 기반 Excel 리포트 생성

    Parameters:
    - results: SAE J2951 결과 딕셔너리
    - test_info: 시험 정보
    - excel_path: Excel 파일 저장 경로
    - chart_path: 레이더 차트 이미지 경로
    - template_path: 템플릿 파일 경로 (None이면 새로 생성)
    """

    # 템플릿이 있으면 복사, 없으면 새로 생성
    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Report"

        # 템플릿 구조 생성
        _create_template_structure(ws)

    # 스타일 정의
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # 값 입력용 셀 스타일
    value_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # 1. Test Report 제목 (A2:F2 병합)
    ws.merge_cells('A2:F2')
    title_cell = ws['A2']
    title_cell.value = "Test Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = center_align

    # 2. 테스트 정보 입력 (좌측 열 - A, 중앙 열 - C, 우측 열 - E)
    # 각 레이블 옆 셀(B, D, F)에 값 입력

    # Row 3-4 (병합된 행)
    ws['B3'] = test_info.get('Test ID', '')
    ws['D3'] = test_info.get('Vehicle ID', '')
    ws['F3'] = test_info.get('Propulsion type', '')

    # Row 5
    ws['B5'] = test_info.get('Test Date', datetime.now().strftime('%Y-%m-%d'))
    ws['D5'] = test_info.get('Vehicle category', '')
    ws['F5'] = test_info.get('Engine ignition type', '')

    # Row 6
    ws['B6'] = test_info.get('Start Time', '')
    ws['D6'] = test_info.get('F0_N', '')  # Road load F₀
    ws['F6'] = test_info.get('Bodywork type', '')

    # Row 7
    ws['B7'] = datetime.now().strftime('%Y-%m-%d')  # Report Date
    ws['D7'] = test_info.get('F1_N_per_kph', '')  # Road load F₁
    ws['F7'] = test_info.get('ICE drive mode', '')

    # Row 8
    ws['B8'] = test_info.get('Supervising person', '')
    ws['D8'] = test_info.get('F2_N_per_kph2', '')  # Road load F₂
    ws['F8'] = test_info.get('PHEV drive mode', '')

    # Row 9
    ws['B9'] = test_info.get('Test cycle', 'WLTC')
    ws['D9'] = test_info.get('Engine displacement', '')
    ws['F9'] = test_info.get('Engine start condition', '')

    # Row 10
    ws['B10'] = test_info.get('Odometer start', '')
    ws['D10'] = test_info.get('Odometer end', '')
    ws['F10'] = test_info.get('Fuel', '')

    # Row 11
    ws['B11'] = test_info.get('Vehicle driver', '')
    ws['D11'] = test_info.get('Mass', '')  # Vehicle test mass
    ws['F11'] = test_info.get('Peak torque', '')

    # Row 12
    ws['B12'] = test_info.get('Vehicle model year', '')
    ws['D12'] = test_info.get('Vehicle gross mass', '')
    ws['F12'] = test_info.get('Wheel drive mode', '')

    # Row 13
    ws['B13'] = test_info.get('Vehicle age', '')
    ws['D13'] = test_info.get('Engine rated power', '')
    ws['F13'] = test_info.get('Air conditioning', '')

    # Row 14
    ws['B14'] = test_info.get('Transmission', '')
    ws['D14'] = test_info.get('Electric motor power', '')
    ws['F14'] = test_info.get('Fuel supply system', '')

    # Row 15
    ws['B15'] = test_info.get('Engine type', '')
    ws['D15'] = test_info.get('Vehicle manufacturer', '')
    ws['F15'] = test_info.get('Vehicle type', '')

    # 3. SAE J2951 report 제목 (A16:F16 병합)
    ws.merge_cells('A16:F16')
    sae_title = ws['A16']
    sae_title.value = "SAE J2951 report"
    sae_title.font = Font(size=14, bold=True)
    sae_title.alignment = center_align

    # 4. SAE J2951 메트릭 헤더 (Row 17)
    headers = ['ER', 'DR', 'EER', 'ASCR', 'IWR', 'RMSSE']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=17, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11)
        cell.alignment = center_align
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 5. SAE J2951 메트릭 값 (Row 18)
    values_row = [
        results.get('ER_pct', 0),
        results.get('DR_pct', 0),
        results.get('EER_pct', 0),
        results.get('ASCR_pct', 0),
        results.get('IWR_pct', 0),
        results.get('RMSSE_mph', 0)
    ]

    for col_idx, value in enumerate(values_row, start=1):
        cell = ws.cell(row=18, column=col_idx)
        cell.value = round(value, 4)
        cell.alignment = center_align
        cell.number_format = '0.0000'

        # 값에 따른 색상 설정
        abs_val = abs(value)
        if abs_val < 1.0:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif abs_val < 2.0:
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 6. DQM 행 추가 (Row 19 병합)
    ws.merge_cells('A19:F19')
    dqm_cell = ws['A19']
    dqm = results.get('DQM', 0)
    dqm_cell.value = f"DQM (Data Quality Metric): {dqm:.4f}"
    dqm_cell.font = Font(bold=True, size=11)
    dqm_cell.alignment = center_align

    # DQM 색상
    if dqm < 1.0:
        dqm_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif dqm < 2.0:
        dqm_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        dqm_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 7. 차트 삽입 (있으면)
    if chart_path and os.path.exists(chart_path):
        try:
            row = 21
            img = XLImage(chart_path)
            img.width = 480
            img.height = 480
            ws.add_image(img, f'A{row}')
        except Exception as e:
            print(f"Warning: Could not insert chart image: {e}")

    # 8. 열 너비 조정
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15

    # 저장
    wb.save(excel_path)
    print(f"✓ Excel report saved: {excel_path}")


def _create_template_structure(ws):
    """템플릿 구조 생성 (템플릿 파일이 없을 때)"""

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    label_font = Font(size=10)

    # Row 1: 빈 행

    # Row 2: Test Report 제목
    ws.merge_cells('A2:F2')

    # Row 3-4: 병합된 행들
    merge_ranges = [
        'A3:A4', 'B3:B4', 'C3:C4', 'D3:D4', 'E3:E4', 'F3:F4'
    ]
    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)

    # 레이블 설정
    labels = {
        'A3': 'Test ID',
        'C3': 'Vehicle ID',
        'E3': 'Propulsion type',
        'A5': 'Test Date',
        'C5': 'Vehicle category',
        'E5': 'Engine ignition type',
        'A6': 'Start Time',
        'C6': 'Road load parameter F₀',
        'E6': 'Bodywork type',
        'A7': 'Report Date',
        'C7': 'Road load parameter F₁',
        'E7': 'ICE drive mode',
        'A8': 'Supervising person',
        'C8': 'Road load parameter F₂',
        'E8': 'PHEV drive mode',
        'A9': 'Test cycle',
        'C9': 'Engine displacement',
        'E9': 'Engine start condition',
        'A10': 'Odometer start value',
        'C10': 'Odometer end value',
        'E10': 'Fuel',
        'A11': 'Vehicle driver',
        'C11': 'Vehicle test mass (kg)',
        'E11': 'Peak torque (Nm)',
        'A12': 'Vehicle model year',
        'C12': 'Vehicle gross mass (kg)',
        'E12': 'Wheel drive mode',
        'A13': 'Vehicle age (months)',
        'C13': 'Engine rated power (kW)',
        'E13': 'Air conditioning used',
        'A14': 'Transmission',
        'C14': 'Electric motor power (kW)',
        'E14': 'Fuel supply system',
        'A15': 'Engine type',
        'C15': 'Vehicle manufacturer',
        'E15': 'Vehicle type',
    }

    for cell_ref, label in labels.items():
        cell = ws[cell_ref]
        cell.value = label
        cell.font = label_font
        cell.alignment = left_align

    # Row 16: SAE J2951 report
    ws.merge_cells('A16:F16')

    # Row 17: 헤더
    # Row 18: 값

    # Row 19: DQM
    ws.merge_cells('A19:F19')


def create_pdf_report(results: Dict[str, float],
                     test_info: Dict[str, str],
                     pdf_path: str):
    """
    PDF 리포트 생성
    """

    with PdfPages(pdf_path) as pdf:
        fig = plt.figure(figsize=(11, 8.5))

        fig.suptitle('SAE J2951 Drive Quality Metrics Report',
                    fontsize=18, fontweight='bold', y=0.98)

        # 시험 정보
        info_text = f"""
Test Information:
  • Test ID: {test_info.get('Test ID', 'N/A')}
  • Test Date: {test_info.get('Test Date', datetime.now().strftime('%Y-%m-%d'))}
  • Vehicle: {test_info.get('Vehicle manufacturer', 'N/A')} {test_info.get('Vehicle type', 'N/A')}
  • Test Mass: {test_info.get('Mass', 'N/A')} kg
  • Test Cycle: {test_info.get('Test cycle', 'WLTC')}
        """
        fig.text(0.1, 0.88, info_text, fontsize=10, verticalalignment='top',
                family='monospace', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.3))

        # 결과 테이블
        ax_table = fig.add_subplot(2, 1, 1)
        ax_table.axis('off')

        dqm = results.get('DQM', 0)
        table_data = [
            ['Metric', 'Value', 'Unit', 'Status'],
            ['ER', f"{results.get('ER_pct', 0):.4f}", '%', ''],
            ['DR', f"{results.get('DR_pct', 0):.4f}", '%', ''],
            ['EER', f"{results.get('EER_pct', 0):.4f}", '%', ''],
            ['ASCR', f"{results.get('ASCR_pct', 0):.4f}", '%', ''],
            ['IWR', f"{results.get('IWR_pct', 0):.4f}", '%', ''],
            ['RMSSE', f"{results.get('RMSSE_mph', 0):.4f}", 'mph', ''],
            ['DQM', f"{dqm:.4f}", '', ''],
        ]

        # Status 열 채우기
        for i in range(1, len(table_data)):
            if i == len(table_data) - 1:  # DQM
                if dqm < 1.0:
                    table_data[i][3] = 'Excellent'
                elif dqm < 2.0:
                    table_data[i][3] = 'Good'
                else:
                    table_data[i][3] = 'Poor'
            else:
                val = float(table_data[i][1])
                if abs(val) < 1.0:
                    table_data[i][3] = 'Good'
                elif abs(val) < 2.0:
                    table_data[i][3] = 'Acceptable'
                else:
                    table_data[i][3] = 'Review'

        table = ax_table.table(cellText=table_data,
                              cellLoc='center',
                              loc='center',
                              colWidths=[0.15, 0.15, 0.1, 0.15])
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1, 2.5)

        # 헤더 스타일
        for i in range(4):
            table[(0, i)].set_facecolor('#366092')
            table[(0, i)].set_text_props(weight='bold', color='white')

        # 값에 따른 색상
        for i in range(1, len(table_data)):
            val = float(table_data[i][1]) if table_data[i][1] else 0
            if i == len(table_data) - 1:  # DQM
                if dqm < 1.0:
                    color = '#C6EFCE'
                elif dqm < 2.0:
                    color = '#FFEB9C'
                else:
                    color = '#FFC7CE'
            else:
                if abs(val) < 1.0:
                    color = '#C6EFCE'
                elif abs(val) < 2.0:
                    color = '#FFEB9C'
                else:
                    color = '#FFC7CE'

            for j in range(4):
                table[(i, j)].set_facecolor(color)

        # 레이더 차트
        ax_radar = fig.add_subplot(2, 1, 2, projection='polar')

        categories = ['ER', 'DR', 'EER', 'ASCR', 'IWR', 'RMSSE']
        values = [
            results.get('ER_pct', 0),
            results.get('DR_pct', 0),
            results.get('EER_pct', 0),
            results.get('ASCR_pct', 0),
            results.get('IWR_pct', 0),
            results.get('RMSSE_mph', 0)
        ]

        angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
        values_plot = values + values[:1]
        angles_plot = angles + angles[:1]

        ax_radar.plot(angles_plot, values_plot, 'o-', linewidth=2.5,
                     color='#DC143C', markersize=8, label='Actual')
        ax_radar.fill(angles_plot, values_plot, alpha=0.3, color='#DC143C')

        zeros = [0] * len(angles_plot)
        ax_radar.plot(angles_plot, zeros, '--', linewidth=2,
                     color='#4169E1', alpha=0.7, label='Target')

        ax_radar.set_xticks(angles)
        ax_radar.set_xticklabels(categories, size=11, weight='bold')

        max_val = max(abs(min(values)), abs(max(values)))
        y_limit = max(3, np.ceil(max_val))
        ax_radar.set_ylim(-y_limit, y_limit)
        yticks = list(range(-int(y_limit), int(y_limit) + 1))
        ax_radar.set_yticks(yticks)
        ax_radar.set_yticklabels([str(y) for y in yticks], size=9)
        ax_radar.grid(True, linestyle='--', alpha=0.4)
        ax_radar.legend(loc='upper right', bbox_to_anchor=(1.25, 1.1))

        plt.tight_layout()
        pdf.savefig(fig, bbox_inches='tight')
        plt.close()

    print(f"✓ PDF report saved: {pdf_path}")


def generate_reports(results: Dict[str, float],
                    test_info: Dict[str, str],
                    output_prefix: str = 'sae_j2951_report',
                    template_path: str = None):
    """
    Excel과 PDF 리포트를 동시에 생성

    Parameters:
    - results: SAE J2951 결과 딕셔너리
    - test_info: 시험 정보
    - output_prefix: 출력 파일명 prefix
    - template_path: Excel 템플릿 파일 경로 (선택사항)

    Returns:
    - dict: 생성된 파일 경로들
    """

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # 파일명 생성
    excel_path = f"{output_prefix}_{timestamp}.xlsx"
    pdf_path = f"{output_prefix}_{timestamp}.pdf"
    chart_path = f"{output_prefix}_{timestamp}_chart.png"

    print("\n" + "="*60)
    print("SAE J2951 Report Generation")
    print("="*60)

    # 1. 레이더 차트 생성
    print("\n[1/3] Generating radar chart...")
    create_radar_chart(
        results.get('ER_pct', 0),
        results.get('DR_pct', 0),
        results.get('EER_pct', 0),
        results.get('ASCR_pct', 0),
        results.get('IWR_pct', 0),
        results.get('RMSSE_mph', 0),
        save_path=chart_path
    )
    print(f"✓ Chart saved: {chart_path}")

    # 2. Excel 리포트 생성 (템플릿 기반)
    print("\n[2/3] Generating Excel report...")
    create_excel_report_from_template(results, test_info, excel_path,
                                     chart_path, template_path)

    # 3. PDF 리포트 생성
    print("\n[3/3] Generating PDF report...")
    create_pdf_report(results, test_info, pdf_path)

    print("\n" + "="*60)
    print("✓ All reports generated successfully!")
    print("="*60)

    return {
        'excel': excel_path,
        'pdf': pdf_path,
        'chart': chart_path
    }


if __name__ == "__main__":
    # 테스트용 예제
    test_results = {
        'ER_pct': -2.345678,
        'DR_pct': 1.234567,
        'EER_pct': -3.456789,
        'ASCR_pct': 0.123456,
        'IWR_pct': -1.987654,
        'RMSSE_mph': 0.456789,
        'DQM': 1.600879
    }

    test_info = {
        'Test ID': 'TEST-2024-001',
        'Vehicle ID': 'VH-12345',
        'Test Date': datetime.now().strftime('%Y-%m-%d'),
        'Mass': '1726.9',
        'F0_N': '35.5',
        'F1_N_per_kph': '1.453',
        'F2_N_per_kph2': '0.03011',
        'Test cycle': 'WLTC',
        'Vehicle manufacturer': 'Hyundai',
        'Vehicle type': 'ACTYON2',
    }

    generate_reports(test_results, test_info)